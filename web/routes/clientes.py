"""
Rutas para gestión de clientes
"""
from io import BytesIO
from flask import Blueprint, render_template, request, redirect, url_for, flash, make_response

from web.auth import admin_required
from src.models import (
    listar_clientes, obtener_cliente, actualizar_cliente,
    crear_cliente, eliminar_cliente, buscar_cliente_por_nombre,
    obtener_estadisticas_clientes
)

clientes_bp = Blueprint('clientes', __name__)


@clientes_bp.route('/')
@admin_required
def listar():
    """Lista todos los clientes con filtros."""
    busqueda = request.args.get('busqueda', '').strip() or None
    con_medidores = request.args.get('con_medidores', '').strip() or None
    filtro_telefono = request.args.get('filtro_telefono', '').strip() or None
    recibe_whatsapp = request.args.get('recibe_whatsapp', '').strip() or None

    clientes = listar_clientes(busqueda=busqueda, con_medidores=con_medidores,
                               filtro_telefono=filtro_telefono, recibe_whatsapp=recibe_whatsapp)

    # Estadísticas
    stats = obtener_estadisticas_clientes(busqueda=busqueda, con_medidores=con_medidores,
                                          filtro_telefono=filtro_telefono)

    # Dict de filtros para chips
    filtros = {
        'busqueda': busqueda,
        'con_medidores': con_medidores,
        'filtro_telefono': filtro_telefono,
        'recibe_whatsapp': recibe_whatsapp
    }

    return render_template('clientes/lista.html',
                           clientes=clientes,
                           stats=stats,
                           filtros=filtros,
                           busqueda=busqueda or '',
                           con_medidores=con_medidores or '',
                           filtro_telefono=filtro_telefono or '',
                           recibe_whatsapp=recibe_whatsapp or '')


@clientes_bp.route('/nuevo', methods=['GET', 'POST'])
@admin_required
def crear():
    """Formulario para crear nuevo cliente."""
    if request.method == 'POST':
        nombre = request.form.get('nombre', '').strip().lower()
        nombre_completo = request.form.get('nombre_completo', '').strip() or None
        rut = request.form.get('rut', '').strip() or None
        telefono = request.form.get('telefono', '').strip() or None
        email = request.form.get('email', '').strip() or None

        if not nombre:
            flash('El nombre es requerido', 'error')
            return redirect(url_for('clientes.crear'))

        # Verificar si ya existe
        if buscar_cliente_por_nombre(nombre):
            flash('Ya existe un cliente con ese nombre', 'error')
            return redirect(url_for('clientes.crear'))

        cliente_id = crear_cliente(nombre, nombre_completo, rut, telefono, email)
        flash('Cliente creado exitosamente', 'success')
        return redirect(url_for('clientes.detalle', cliente_id=cliente_id))

    return render_template('clientes/crear.html')


@clientes_bp.route('/<int:cliente_id>')
@admin_required
def detalle(cliente_id):
    """Muestra detalle de un cliente."""
    cliente = obtener_cliente(cliente_id)
    if not cliente:
        flash('Cliente no encontrado', 'error')
        return redirect(url_for('clientes.listar'))

    return render_template('clientes/detalle.html', cliente=cliente)


@clientes_bp.route('/<int:cliente_id>/editar', methods=['GET', 'POST'])
@admin_required
def editar(cliente_id):
    """Edita datos de un cliente."""
    cliente = obtener_cliente(cliente_id)
    if not cliente:
        flash('Cliente no encontrado', 'error')
        return redirect(url_for('clientes.listar'))

    if request.method == 'POST':
        nombre = request.form.get('nombre', '').strip().lower() or None
        nombre_completo = request.form.get('nombre_completo', '').strip() or None
        rut = request.form.get('rut', '').strip() or None
        telefono = request.form.get('telefono', '').strip() or None
        email = request.form.get('email', '').strip() or None
        recibe_boleta_whatsapp = 1 if request.form.get('recibe_boleta_whatsapp') == 'on' else 0
        volver_a_lista = request.form.get('volver_a_lista') == '1'

        # Validar que si activa recibe_boleta_whatsapp, debe tener telefono
        if recibe_boleta_whatsapp == 1 and not telefono:
            flash('Para recibir boleta por WhatsApp debe tener un numero de telefono', 'error')
            if volver_a_lista:
                return redirect(url_for('clientes.listar'))
            return redirect(url_for('clientes.editar', cliente_id=cliente_id))

        # Verificar si el nuevo nombre ya existe (y no es el mismo cliente)
        if nombre and nombre != cliente['nombre']:
            existente = buscar_cliente_por_nombre(nombre)
            if existente and existente['id'] != cliente_id:
                flash('Ya existe otro cliente con ese nombre', 'error')
                if volver_a_lista:
                    return redirect(url_for('clientes.listar'))
                return redirect(url_for('clientes.editar', cliente_id=cliente_id))

        actualizar_cliente(cliente_id, nombre=nombre, nombre_completo=nombre_completo,
                          rut=rut, telefono=telefono, email=email,
                          recibe_boleta_whatsapp=recibe_boleta_whatsapp)
        flash('Cliente actualizado', 'success')

        # Si viene del modal en la lista, volver a la lista con filtros
        if volver_a_lista:
            filtros = {
                'busqueda': request.form.get('busqueda', '').strip() or None,
                'con_medidores': request.form.get('con_medidores', '').strip() or None,
                'filtro_telefono': request.form.get('filtro_telefono', '').strip() or None
            }
            # Limpiar valores None
            filtros = {k: v for k, v in filtros.items() if v}
            return redirect(url_for('clientes.listar', **filtros))

        return redirect(url_for('clientes.detalle', cliente_id=cliente_id))

    return render_template('clientes/editar.html', cliente=cliente)


@clientes_bp.route('/<int:cliente_id>/eliminar', methods=['POST'])
@admin_required
def eliminar(cliente_id):
    """Elimina un cliente."""
    cliente = obtener_cliente(cliente_id)
    if not cliente:
        flash('Cliente no encontrado', 'error')
        return redirect(url_for('clientes.listar'))

    exito, motivo = eliminar_cliente(cliente_id)

    if exito:
        flash('Cliente eliminado exitosamente', 'success')
    elif motivo == "medidores":
        flash('No se puede eliminar el cliente porque tiene medidores asociados', 'error')
    else:
        flash('Error al eliminar el cliente', 'error')

    return redirect(url_for('clientes.listar'))


@clientes_bp.route('/exportar')
@admin_required
def exportar():
    """Exporta clientes filtrados a Excel con columnas seleccionables."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from datetime import datetime

    # Obtener parametros de filtro
    busqueda = request.args.get('busqueda', '').strip() or None
    con_medidores = request.args.get('con_medidores', '').strip() or None
    filtro_telefono = request.args.get('filtro_telefono', '').strip() or None

    # Obtener columnas seleccionadas (por defecto todas)
    columnas_param = request.args.get('columnas', '').strip()
    if columnas_param:
        columnas_seleccionadas = columnas_param.split(',')
    else:
        columnas_seleccionadas = ['id', 'nombre', 'nombre_completo', 'rut', 'telefono', 'email', 'medidores']

    # Definicion de columnas disponibles
    columnas_config = {
        'id': {'header': 'ID', 'field': 'id', 'width': 8},
        'nombre': {'header': 'Nombre', 'field': 'nombre', 'width': 25},
        'nombre_completo': {'header': 'Nombre Completo', 'field': 'nombre_completo', 'width': 30},
        'rut': {'header': 'RUT', 'field': 'rut', 'width': 15},
        'telefono': {'header': 'Telefono', 'field': 'telefono', 'width': 18},
        'email': {'header': 'Email', 'field': 'email', 'width': 30},
        'medidores': {'header': 'Medidores', 'field': 'num_medidores', 'width': 12}
    }

    # Filtrar solo columnas validas
    columnas = [c for c in columnas_seleccionadas if c in columnas_config]
    if not columnas:
        columnas = ['id', 'nombre']

    num_cols = len(columnas)

    # Obtener clientes con filtros
    clientes = listar_clientes(busqueda=busqueda, con_medidores=con_medidores,
                               filtro_telefono=filtro_telefono)

    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Clientes"

    # Estilos
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Título y fecha
    last_col_letter = get_column_letter(num_cols)
    ws.merge_cells(f'A1:{last_col_letter}1')
    ws['A1'] = 'LISTADO DE CLIENTES'
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal="center")

    ws.merge_cells(f'A2:{last_col_letter}2')
    ws['A2'] = f'Generado: {datetime.now().strftime("%d/%m/%Y %H:%M")}'
    ws['A2'].alignment = Alignment(horizontal="center")

    # Filtros aplicados
    row = 3
    if any([busqueda, con_medidores, filtro_telefono]):
        ws.merge_cells(f'A{row}:{last_col_letter}{row}')
        filtros_texto = []
        if busqueda:
            filtros_texto.append(f'Busqueda: {busqueda}')
        if con_medidores == 'si':
            filtros_texto.append('Con medidores')
        elif con_medidores == 'no':
            filtros_texto.append('Sin medidores')
        if filtro_telefono == 'sin':
            filtros_texto.append('Sin telefono')
        elif filtro_telefono == 'con':
            filtros_texto.append('Con telefono')

        ws[f'A{row}'] = 'Filtros aplicados: ' + ' | '.join(filtros_texto)
        ws[f'A{row}'].font = Font(italic=True)
        ws[f'A{row}'].alignment = Alignment(horizontal="center")
        row += 1

    # Resumen
    row += 1
    ws.merge_cells(f'A{row}:{last_col_letter}{row}')
    ws[f'A{row}'] = f'Total: {len(clientes)} cliente(s)'
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'A{row}'].alignment = Alignment(horizontal="center")
    ws[f'A{row}'].fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")

    # Espacio
    row += 2

    # Encabezados de tabla (solo columnas seleccionadas)
    for col_idx, col_key in enumerate(columnas, 1):
        cell = ws.cell(row=row, column=col_idx, value=columnas_config[col_key]['header'])
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border

    # Datos (solo columnas seleccionadas)
    row += 1
    for cliente in clientes:
        for col_idx, col_key in enumerate(columnas, 1):
            field = columnas_config[col_key]['field']
            valor = cliente.get(field, '')
            if valor is None or valor == '':
                valor = '-' if col_key != 'medidores' else 0
            cell = ws.cell(row=row, column=col_idx, value=valor)
            cell.border = border
        row += 1

    # Ajustar anchos de columna
    for col_idx, col_key in enumerate(columnas, 1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = columnas_config[col_key]['width']

    # Preparar respuesta
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Nombre del archivo con timestamp
    filename = f'clientes_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'

    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename={filename}'

    return response
