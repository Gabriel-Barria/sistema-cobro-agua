"""
Rutas para gestion de boletas - Sistema desacoplado
"""
import os
from io import BytesIO
from flask import Blueprint, render_template, request, redirect, url_for, flash, send_file, make_response
from werkzeug.utils import secure_filename
from weasyprint import HTML

from web.auth import admin_required, get_current_user
from src.models_boletas import (
    obtener_configuracion, guardar_configuracion,
    crear_boleta, obtener_boleta, obtener_boleta_por_lectura,
    listar_boletas, desmarcar_boleta_pagada,
    guardar_comprobante, eliminar_boleta,
    obtener_lectura_anterior, calcular_consumo,
    obtener_lecturas_sin_boleta, obtener_años_disponibles,
    obtener_estadisticas_boletas
)
from src.models import listar_clientes, listar_medidores, obtener_lectura
from src.models_pagos import (
    listar_pagos, obtener_pago, aprobar_pago, rechazar_pago,
    registrar_pago_directo, listar_saldos_clientes, ajustar_saldo_cliente,
    obtener_resumen_cuenta_cliente, obtener_saldo_cliente
)

boletas_bp = Blueprint('boletas', __name__)

# Directorio para comprobantes
BASE_DIR = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
COMPROBANTES_DIR = os.path.join(BASE_DIR, 'comprobantes')

ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif', 'pdf'}


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


# =============================================================================
# CONFIGURACION
# =============================================================================

@boletas_bp.route('/configuracion', methods=['GET', 'POST'])
@admin_required
def configuracion():
    """Configuracion de tarifas para boletas."""
    if request.method == 'POST':
        try:
            cargo_fijo = float(request.form.get('cargo_fijo', 0))
            precio_m3 = float(request.form.get('precio_m3', 0))

            guardar_configuracion(cargo_fijo, precio_m3)
            flash('Configuracion guardada exitosamente', 'success')
        except ValueError:
            flash('Los valores deben ser numericos', 'error')

        return redirect(url_for('boletas.configuracion'))

    config = obtener_configuracion()
    return render_template('boletas/configuracion.html', config=config)


# =============================================================================
# HISTORIAL DE PAGOS (LEGACY - REDIRIGE A NUEVO SISTEMA)
# =============================================================================

@boletas_bp.route('/historial-pagos')
@admin_required
def historial_pagos():
    """Redirige al nuevo sistema de pagos."""
    return redirect(url_for('boletas.pagos_lista'))


# =============================================================================
# LISTADO Y FILTROS
# =============================================================================

@boletas_bp.route('/')
@admin_required
def listar():
    """Lista boletas con filtros."""
    # Obtener parametros de filtro
    cliente_id = request.args.get('cliente_id', type=int)
    medidor_id = request.args.get('medidor_id', type=int)
    pagada = request.args.get('pagada', type=int)
    sin_comprobante = request.args.get('sin_comprobante', type=int) == 1
    año = request.args.get('año', type=int)
    mes = request.args.get('mes', type=int)

    # Obtener parametros de ordenamiento
    sort_by = request.args.get('sort_by', type=str)
    sort_order = request.args.get('sort_order', type=str)

    boletas = listar_boletas(
        cliente_id=cliente_id,
        medidor_id=medidor_id,
        pagada=pagada,
        sin_comprobante=sin_comprobante,
        año=año,
        mes=mes
    )

    # Datos para filtros
    clientes = listar_clientes()
    años = obtener_años_disponibles()
    # Estadísticas con los mismos filtros aplicados
    stats = obtener_estadisticas_boletas(
        cliente_id=cliente_id,
        medidor_id=medidor_id,
        pagada=pagada,
        sin_comprobante=sin_comprobante,
        año=año,
        mes=mes
    )

    # Medidores del cliente seleccionado
    medidores = []
    if cliente_id:
        medidores = listar_medidores(cliente_id)

    return render_template('boletas/lista.html',
                           boletas=boletas,
                           clientes=clientes,
                           medidores=medidores,
                           años=años,
                           stats=stats,
                           filtros={
                               'cliente_id': cliente_id,
                               'medidor_id': medidor_id,
                               'pagada': pagada,
                               'sin_comprobante': sin_comprobante,
                               'año': año,
                               'mes': mes,
                               'sort_by': sort_by,
                               'sort_order': sort_order
                           })


# =============================================================================
# DETALLE DE BOLETA
# =============================================================================

@boletas_bp.route('/<int:boleta_id>')
@admin_required
def detalle(boleta_id):
    """Muestra detalle de una boleta."""
    from src.database import get_connection

    boleta = obtener_boleta(boleta_id)
    if not boleta:
        flash('Boleta no encontrada', 'error')
        return redirect(url_for('boletas.listar'))

    # Obtener pagos asociados a esta boleta
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute('''
        SELECT p.id, p.numero_pago, p.monto_total, p.estado, p.comprobante_path,
               p.fecha_envio, p.fecha_procesamiento, p.motivo_rechazo,
               p.metodo_pago, p.created_at, pb.monto_aplicado
        FROM pagos p
        JOIN pago_boletas pb ON p.id = pb.pago_id
        WHERE pb.boleta_id = %s
        ORDER BY p.created_at DESC
    ''', (boleta_id,))
    pagos = [dict(p) for p in cursor.fetchall()]
    conn.close()

    return render_template('boletas/detalle.html',
                          boleta=boleta,
                          pagos=pagos)


# =============================================================================
# CREAR BOLETA INDIVIDUAL
# =============================================================================

@boletas_bp.route('/crear', methods=['GET', 'POST'])
@admin_required
def crear():
    """Formulario para crear una boleta desde una lectura."""
    config = obtener_configuracion()
    if not config:
        flash('Debe configurar las tarifas antes de crear boletas', 'error')
        return redirect(url_for('boletas.configuracion'))

    if request.method == 'POST':
        lectura_id = request.form.get('lectura_id', type=int)

        if not lectura_id:
            flash('Debe seleccionar una lectura', 'error')
            return redirect(url_for('boletas.crear'))

        # Verificar que no exista boleta para esta lectura
        if obtener_boleta_por_lectura(lectura_id):
            flash('Ya existe una boleta para esta lectura', 'error')
            return redirect(url_for('boletas.crear'))

        # Obtener datos de la lectura
        lectura = obtener_lectura(lectura_id)
        if not lectura:
            flash('Lectura no encontrada', 'error')
            return redirect(url_for('boletas.crear'))

        # Calcular consumo
        lectura_anterior = obtener_lectura_anterior(
            lectura['medidor_id'],
            lectura['año'],
            lectura['mes']
        )
        consumo = calcular_consumo(lectura['lectura_m3'], lectura_anterior)

        # Crear boleta
        boleta_id = crear_boleta(
            lectura_id=lectura_id,
            cliente_nombre=lectura['cliente_nombre'],
            medidor_id=lectura['medidor_id'],
            periodo_año=lectura['año'],
            periodo_mes=lectura['mes'],
            lectura_actual=lectura['lectura_m3'],
            lectura_anterior=lectura_anterior,
            consumo_m3=consumo,
            cargo_fijo=config['cargo_fijo'],
            precio_m3=config['precio_m3']
        )

        flash('Boleta creada exitosamente', 'success')
        return redirect(url_for('boletas.detalle', boleta_id=boleta_id))

    # GET: Mostrar formulario
    cliente_id = request.args.get('cliente_id', type=int)
    año = request.args.get('año', type=int)
    mes = request.args.get('mes', type=int)

    lecturas = obtener_lecturas_sin_boleta(año=año, mes=mes, cliente_id=cliente_id)
    clientes = listar_clientes()
    años = obtener_años_disponibles()

    return render_template('boletas/crear.html',
                           lecturas=lecturas,
                           clientes=clientes,
                           años=años,
                           config=config,
                           filtros={
                               'cliente_id': cliente_id,
                               'año': año,
                               'mes': mes
                           })


# =============================================================================
# CREAR BOLETAS MASIVAS
# =============================================================================

@boletas_bp.route('/crear-masivo', methods=['GET', 'POST'])
@admin_required
def crear_masivo():
    """Crear multiples boletas a la vez."""
    config = obtener_configuracion()
    if not config:
        flash('Debe configurar las tarifas antes de crear boletas', 'error')
        return redirect(url_for('boletas.configuracion'))

    if request.method == 'POST':
        lectura_ids = request.form.getlist('lecturas')

        if not lectura_ids:
            flash('Debe seleccionar al menos una lectura', 'error')
            return redirect(url_for('boletas.crear_masivo'))

        creadas = 0
        omitidas = 0

        for lectura_id in lectura_ids:
            lectura_id = int(lectura_id)

            # Verificar que no exista boleta
            if obtener_boleta_por_lectura(lectura_id):
                omitidas += 1
                continue

            # Obtener datos de la lectura
            lectura = obtener_lectura(lectura_id)
            if not lectura:
                omitidas += 1
                continue

            # Calcular consumo
            lectura_anterior = obtener_lectura_anterior(
                lectura['medidor_id'],
                lectura['año'],
                lectura['mes']
            )
            consumo = calcular_consumo(lectura['lectura_m3'], lectura_anterior)

            # Crear boleta
            crear_boleta(
                lectura_id=lectura_id,
                cliente_nombre=lectura['cliente_nombre'],
                medidor_id=lectura['medidor_id'],
                periodo_año=lectura['año'],
                periodo_mes=lectura['mes'],
                lectura_actual=lectura['lectura_m3'],
                lectura_anterior=lectura_anterior,
                consumo_m3=consumo,
                cargo_fijo=config['cargo_fijo'],
                precio_m3=config['precio_m3']
            )
            creadas += 1

        if creadas > 0:
            flash(f'{creadas} boletas creadas exitosamente', 'success')
        if omitidas > 0:
            flash(f'{omitidas} lecturas omitidas (ya tenian boleta)', 'warning')

        return redirect(url_for('boletas.listar'))

    # GET: Mostrar formulario
    cliente_id = request.args.get('cliente_id', type=int)
    año = request.args.get('año', type=int)
    mes = request.args.get('mes', type=int)

    lecturas = obtener_lecturas_sin_boleta(año=año, mes=mes, cliente_id=cliente_id)
    clientes = listar_clientes()
    años = obtener_años_disponibles()

    return render_template('boletas/crear_masivo.html',
                           lecturas=lecturas,
                           clientes=clientes,
                           años=años,
                           config=config,
                           filtros={
                               'cliente_id': cliente_id,
                               'año': año,
                               'mes': mes
                           })


# =============================================================================
# MARCAR COMO PAGADA
# =============================================================================

@boletas_bp.route('/<int:boleta_id>/marcar-pagada', methods=['POST'])
@admin_required
def marcar_pagada(boleta_id):
    """Marca una boleta como pagada registrando un pago directo."""
    from decimal import Decimal
    from src.database import get_connection

    boleta = obtener_boleta(boleta_id)
    if not boleta:
        flash('Boleta no encontrada', 'error')
        return redirect(url_for('boletas.listar'))

    metodo_pago = request.form.get('metodo_pago', 'efectivo')

    # Obtener cliente_id desde el medidor
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute('SELECT cliente_id FROM medidores WHERE id = %s', (boleta['medidor_id'],))
    medidor = cursor.fetchone()
    conn.close()

    if not medidor:
        flash('Error: medidor no encontrado', 'error')
        return redirect(url_for('boletas.detalle', boleta_id=boleta_id))

    usuario = get_current_user()
    usuario_id = usuario['id'] if usuario else None

    # Calcular monto (saldo pendiente o total)
    monto = Decimal(str(boleta.get('saldo_pendiente', boleta['total'])))

    try:
        resultado = registrar_pago_directo(
            cliente_id=medidor['cliente_id'],
            monto_total=monto,
            boletas_ids=[boleta_id],
            metodo_pago=metodo_pago,
            usuario_id=usuario_id
        )
        flash(f'Pago {resultado["numero_pago"]} registrado. Boleta marcada como pagada.', 'success')
    except Exception as e:
        flash(f'Error al registrar pago: {str(e)}', 'error')

    # Verificar si se solicitó volver al listado
    volver_a_lista = request.form.get('volver_a_lista', False)
    referer = request.referrer or ''

    if volver_a_lista or (referer and referer.rstrip('/').endswith('/boletas')):
        # Preservar filtros al redirigir
        return redirect(construir_url_con_filtros('boletas.listar'))

    # Por defecto, volver al detalle (comportamiento original)
    return redirect(url_for('boletas.detalle', boleta_id=boleta_id))


@boletas_bp.route('/<int:boleta_id>/desmarcar-pagada', methods=['POST'])
@admin_required
def desmarcar_pagada(boleta_id):
    """Desmarca una boleta como pagada."""
    boleta = obtener_boleta(boleta_id)
    if not boleta:
        flash('Boleta no encontrada', 'error')
        return redirect(url_for('boletas.listar'))

    if desmarcar_boleta_pagada(boleta_id):
        flash('Boleta desmarcada', 'success')
    else:
        flash('Error al desmarcar boleta', 'error')

    # Verificar si se solicitó volver al listado
    volver_a_lista = request.form.get('volver_a_lista', False)
    referer = request.referrer or ''

    if volver_a_lista or (referer and referer.rstrip('/').endswith('/boletas')):
        # Preservar filtros al redirigir
        return redirect(construir_url_con_filtros('boletas.listar'))

    # Por defecto, volver al detalle (comportamiento original)
    return redirect(url_for('boletas.detalle', boleta_id=boleta_id))


# =============================================================================
# APROBAR / RECHAZAR BOLETAS (WORKFLOW DE REVISION)
# =============================================================================

@boletas_bp.route('/<int:boleta_id>/aprobar', methods=['POST'])
@admin_required
def aprobar(boleta_id):
    """Aprueba el pago asociado a la boleta: En Revisión (1) → Pagada (2)"""
    from src.database import get_connection

    boleta = obtener_boleta(boleta_id)
    if not boleta:
        flash('Boleta no encontrada', 'error')
        return redirect(url_for('boletas.listar'))

    if boleta['pagada'] != 1:
        flash('Solo se pueden aprobar boletas en estado "En Revisión"', 'error')
        return redirect(url_for('boletas.listar'))

    # Buscar el pago asociado
    pago_id = _obtener_o_crear_pago_para_boleta(boleta)

    usuario = get_current_user()
    usuario_id = usuario['id'] if usuario else None

    exito, mensaje = aprobar_pago(pago_id, usuario_id)
    if exito:
        flash(mensaje, 'success')
    else:
        flash(mensaje, 'error')

    return redirect(construir_url_con_filtros('boletas.listar'))


@boletas_bp.route('/<int:boleta_id>/rechazar', methods=['POST'])
@admin_required
def rechazar(boleta_id):
    """Rechaza el pago asociado a la boleta: En Revisión (1) → Pendiente (0)"""
    from src.database import get_connection

    boleta = obtener_boleta(boleta_id)
    if not boleta:
        flash('Boleta no encontrada', 'error')
        return redirect(url_for('boletas.listar'))

    if boleta['pagada'] != 1:
        flash('Solo se pueden rechazar boletas en estado "En Revisión"', 'error')
        return redirect(url_for('boletas.listar'))

    motivo = request.form.get('motivo', '').strip()
    if not motivo:
        flash('Debe proporcionar un motivo de rechazo', 'error')
        return redirect(url_for('boletas.listar'))

    # Buscar el pago asociado
    pago_id = _obtener_o_crear_pago_para_boleta(boleta)

    usuario = get_current_user()
    usuario_id = usuario['id'] if usuario else None

    exito, mensaje = rechazar_pago(pago_id, motivo, usuario_id)
    if exito:
        flash(f'{mensaje}. Motivo: {motivo}', 'success')
    else:
        flash(mensaje, 'error')

    return redirect(construir_url_con_filtros('boletas.listar'))


def _obtener_o_crear_pago_para_boleta(boleta):
    """
    Busca el pago en_revision asociado a la boleta.
    Si no existe (dato huerfano de migracion), crea uno.
    Tambien busca otras boletas con el mismo comprobante.
    """
    from src.database import get_connection
    from src.models_pagos import generar_numero_pago
    from decimal import Decimal

    boleta_id = boleta['id']

    conn = get_connection()
    cursor = conn.cursor()

    # Buscar pago existente
    cursor.execute('''
        SELECT p.id as pago_id
        FROM pagos p
        JOIN pago_boletas pb ON p.id = pb.pago_id
        WHERE pb.boleta_id = %s AND p.estado = 'en_revision'
        LIMIT 1
    ''', (boleta_id,))
    pago_result = cursor.fetchone()

    if pago_result:
        conn.close()
        return pago_result['pago_id']

    # No existe pago - crear uno para esta boleta y otras con mismo comprobante
    comprobante_path = boleta.get('comprobante_path')

    # Buscar todas las boletas en revision con el mismo comprobante
    if comprobante_path:
        cursor.execute('''
            SELECT b.id, b.total, b.saldo_pendiente
            FROM boletas b
            WHERE b.comprobante_path = %s AND b.pagada = 1
        ''', (comprobante_path,))
    else:
        cursor.execute('''
            SELECT b.id, b.total, b.saldo_pendiente
            FROM boletas b
            WHERE b.id = %s AND b.pagada = 1
        ''', (boleta_id,))

    boletas_relacionadas = cursor.fetchall()

    # Obtener cliente_id
    cursor.execute('SELECT cliente_id FROM medidores WHERE id = %s', (boleta['medidor_id'],))
    medidor = cursor.fetchone()
    cliente_id = medidor['cliente_id']

    # Calcular monto total
    monto_total = sum(Decimal(str(b['saldo_pendiente'] or b['total'])) for b in boletas_relacionadas)

    # Crear pago
    numero_pago = generar_numero_pago()
    cursor.execute('''
        INSERT INTO pagos (numero_pago, cliente_id, monto_total, monto_aplicado,
                          comprobante_path, metodo_pago, estado, fecha_envio)
        VALUES (%s, %s, %s, %s, %s, 'transferencia', 'en_revision', CURRENT_DATE)
        RETURNING id
    ''', (numero_pago, cliente_id, monto_total, monto_total, comprobante_path))
    pago_id = cursor.fetchone()['id']

    # Crear relaciones pago_boletas
    for b in boletas_relacionadas:
        monto_boleta = Decimal(str(b['saldo_pendiente'] or b['total']))
        cursor.execute('''
            INSERT INTO pago_boletas (pago_id, boleta_id, monto_aplicado, es_pago_completo)
            VALUES (%s, %s, %s, TRUE)
        ''', (pago_id, b['id'], monto_boleta))

    conn.commit()
    conn.close()

    return pago_id


# =============================================================================
# SUBIR COMPROBANTE
# =============================================================================

@boletas_bp.route('/<int:boleta_id>/subir-comprobante', methods=['POST'])
@admin_required
def subir_comprobante(boleta_id):
    """Sube un comprobante de pago."""
    boleta = obtener_boleta(boleta_id)
    if not boleta:
        flash('Boleta no encontrada', 'error')
        return redirect(url_for('boletas.listar'))

    if 'comprobante' not in request.files:
        flash('No se selecciono archivo', 'error')
        # Verificar si viene del listado para preservar filtros
        volver_a_lista = request.form.get('volver_a_lista', False)
        if volver_a_lista:
            return redirect(construir_url_con_filtros('boletas.listar'))
        return redirect(url_for('boletas.detalle', boleta_id=boleta_id))

    file = request.files['comprobante']
    if file.filename == '':
        flash('No se selecciono archivo', 'error')
        # Verificar si viene del listado para preservar filtros
        volver_a_lista = request.form.get('volver_a_lista', False)
        if volver_a_lista:
            return redirect(construir_url_con_filtros('boletas.listar'))
        return redirect(url_for('boletas.detalle', boleta_id=boleta_id))

    if file and allowed_file(file.filename):
        filename = secure_filename(file.filename)
        # Crear directorio para la boleta
        boleta_dir = os.path.join(COMPROBANTES_DIR, f'boleta_{boleta_id}')
        os.makedirs(boleta_dir, exist_ok=True)

        # Guardar archivo
        filepath = os.path.join(boleta_dir, filename)
        file.save(filepath)

        # Guardar ruta relativa en BD
        comprobante_path = f'comprobantes/boleta_{boleta_id}/{filename}'
        guardar_comprobante(boleta_id, comprobante_path)

        flash('Comprobante subido exitosamente', 'success')
    else:
        flash('Tipo de archivo no permitido', 'error')

    # Verificar si se solicitó volver al listado
    volver_a_lista = request.form.get('volver_a_lista', False)
    referer = request.referrer or ''

    if volver_a_lista or (referer and referer.rstrip('/').endswith('/boletas')):
        # Preservar filtros al redirigir
        return redirect(construir_url_con_filtros('boletas.listar'))

    # Por defecto, volver al detalle (comportamiento original)
    return redirect(url_for('boletas.detalle', boleta_id=boleta_id))


# =============================================================================
# DESCARGAR BOLETA
# =============================================================================

@boletas_bp.route('/<int:boleta_id>/descargar')
@admin_required
def descargar(boleta_id):
    """Descarga la boleta en formato PDF."""
    boleta = obtener_boleta(boleta_id)
    if not boleta:
        flash('Boleta no encontrada', 'error')
        return redirect(url_for('boletas.listar'))

    # Nombres de meses en espanol
    meses = {
        1: 'Enero', 2: 'Febrero', 3: 'Marzo', 4: 'Abril',
        5: 'Mayo', 6: 'Junio', 7: 'Julio', 8: 'Agosto',
        9: 'Septiembre', 10: 'Octubre', 11: 'Noviembre', 12: 'Diciembre'
    }

    # Obtener la lectura actual asociada para la foto del medidor y fecha
    lectura_actual = None
    foto_lectura = None
    fecha_lectura_actual = None
    if boleta['lectura_id']:
        lectura_actual = obtener_lectura(boleta['lectura_id'])
        if lectura_actual:
            if lectura_actual.get('foto_path') and lectura_actual.get('foto_path') != '':
                # Convertir ruta relativa a absoluta para WeasyPrint
                # foto_path viene como 'medidor_X/...' sin prefijo de carpeta
                foto_lectura = os.path.join(BASE_DIR, 'fotos', lectura_actual['foto_path'])
            fecha_lectura_actual = lectura_actual.get('fecha_lectura')

    # Obtener la lectura anterior y su fecha
    fecha_lectura_anterior = None
    if boleta['lectura_anterior'] is not None:
        # Buscar la lectura anterior por medidor_id y periodo
        medidor_id = boleta['medidor_id']
        año = boleta['periodo_año']
        mes = boleta['periodo_mes']

        # Calcular periodo anterior
        if mes == 1:
            mes_anterior = 12
            año_anterior = año - 1
        else:
            mes_anterior = mes - 1
            año_anterior = año

        # Buscar lectura anterior
        from src.database import get_connection
        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute('''
            SELECT fecha_lectura FROM lecturas
            WHERE medidor_id = %s AND año = %s AND mes = %s
        ''', (medidor_id, año_anterior, mes_anterior))
        lectura_ant = cursor.fetchone()
        conn.close()

        if lectura_ant:
            fecha_lectura_anterior = lectura_ant['fecha_lectura']

    # Renderizar template HTML
    html_string = render_template('boletas/boleta_pdf.html',
                                   boleta=boleta,
                                   meses=meses,
                                   foto_lectura=foto_lectura,
                                   fecha_lectura_actual=fecha_lectura_actual,
                                   fecha_lectura_anterior=fecha_lectura_anterior)

    # Generar PDF usando WeasyPrint
    pdf_file = BytesIO()
    HTML(string=html_string, base_url=BASE_DIR).write_pdf(pdf_file)
    pdf_file.seek(0)

    # Devolver PDF como descarga
    return send_file(
        pdf_file,
        mimetype='application/pdf',
        as_attachment=True,
        download_name=f'boleta_{boleta["numero_boleta"]}.pdf'
    )


# =============================================================================
# ELIMINAR BOLETA
# =============================================================================

@boletas_bp.route('/<int:boleta_id>/eliminar', methods=['POST'])
@admin_required
def eliminar(boleta_id):
    """Elimina una boleta."""
    boleta = obtener_boleta(boleta_id)
    if not boleta:
        flash('Boleta no encontrada', 'error')
        return redirect(url_for('boletas.listar'))

    if eliminar_boleta(boleta_id):
        flash('Boleta eliminada', 'success')
    else:
        flash('Error al eliminar boleta', 'error')

    return redirect(url_for('boletas.listar'))


# =============================================================================
# EXPORTAR A EXCEL
# =============================================================================

@boletas_bp.route('/exportar')
@admin_required
def exportar():
    """Exporta boletas filtradas a Excel."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from datetime import datetime

    # Obtener parametros de filtro (los mismos que listar)
    cliente_id = request.args.get('cliente_id', type=int)
    medidor_id = request.args.get('medidor_id', type=int)
    pagada = request.args.get('pagada', type=int)
    sin_comprobante = request.args.get('sin_comprobante', type=int) == 1
    año = request.args.get('año', type=int)
    mes = request.args.get('mes', type=int)

    # Obtener boletas con filtros
    boletas = listar_boletas(
        cliente_id=cliente_id,
        medidor_id=medidor_id,
        pagada=pagada,
        sin_comprobante=sin_comprobante,
        año=año,
        mes=mes
    )

    # Obtener estadísticas con filtros
    stats = obtener_estadisticas_boletas(
        cliente_id=cliente_id,
        medidor_id=medidor_id,
        pagada=pagada,
        sin_comprobante=sin_comprobante,
        año=año,
        mes=mes
    )

    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Boletas"

    # Estilos
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Título y fecha
    ws.merge_cells('A1:H1')
    ws['A1'] = 'REPORTE DE BOLETAS'
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal="center")

    ws.merge_cells('A2:H2')
    ws['A2'] = f'Generado: {datetime.now().strftime("%d/%m/%Y %H:%M")}'
    ws['A2'].alignment = Alignment(horizontal="center")

    # Filtros aplicados (si existen)
    row = 3
    if any([cliente_id, medidor_id, pagada is not None, sin_comprobante, año, mes]):
        ws.merge_cells(f'A{row}:H{row}')
        filtros_texto = []
        if año:
            filtros_texto.append(f'Año: {año}')
        if mes:
            filtros_texto.append(f'Mes: {mes}')
        if pagada == 0:
            filtros_texto.append('Estado: Pendientes')
        elif pagada == 1:
            filtros_texto.append('Estado: Pagadas')
        if sin_comprobante:
            filtros_texto.append('Sin comprobante')

        ws[f'A{row}'] = 'Filtros aplicados: ' + ' | '.join(filtros_texto)
        ws[f'A{row}'].font = Font(italic=True)
        ws[f'A{row}'].alignment = Alignment(horizontal="center")
        row += 1

    # Estadísticas
    row += 1
    ws.merge_cells(f'A{row}:H{row}')
    ws[f'A{row}'] = 'RESUMEN'
    ws[f'A{row}'].font = Font(bold=True, size=12)
    ws[f'A{row}'].alignment = Alignment(horizontal="center")
    ws[f'A{row}'].fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")

    row += 1
    stats_data = [
        ['Total Boletas:', stats.get('total', 0)],
        ['Pagadas:', stats.get('pagadas', 0)],
        ['Pendientes:', stats.get('pendientes', 0)],
        ['Sin Comprobante:', stats.get('sin_comprobante', 0)],
        ['Monto Por Cobrar:', f"${stats.get('monto_pendiente', 0):,.0f}"]
    ]

    for label, value in stats_data:
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = value
        row += 1

    # Espacio
    row += 1

    # Encabezados de tabla
    headers = ['N° Boleta', 'Cliente', 'Medidor', 'Período', 'Consumo (m³)', 'Total', 'Estado', 'Comprobante']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border

    # Datos
    row += 1
    for boleta in boletas:
        ws.cell(row=row, column=1, value=boleta.get('numero_boleta', '')).border = border
        ws.cell(row=row, column=2, value=boleta.get('cliente_nombre', '')).border = border
        ws.cell(row=row, column=3, value=boleta.get('numero_medidor', '')).border = border
        ws.cell(row=row, column=4, value=f"{boleta.get('periodo_mes', '')}/{boleta.get('periodo_año', '')}").border = border
        ws.cell(row=row, column=5, value=boleta.get('consumo_m3', 0)).border = border

        total_cell = ws.cell(row=row, column=6, value=boleta.get('total', 0))
        total_cell.border = border
        total_cell.number_format = '$#,##0'

        estado = 'Pagada' if boleta.get('pagada') == 1 else 'Pendiente'
        estado_cell = ws.cell(row=row, column=7, value=estado)
        estado_cell.border = border
        if boleta.get('pagada') == 1:
            estado_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        else:
            estado_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        comprobante = 'Sí' if boleta.get('comprobante_path') else 'No'
        ws.cell(row=row, column=8, value=comprobante).border = border

        row += 1

    # Ajustar anchos de columna
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 15

    # Preparar respuesta
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Nombre del archivo con timestamp
    filename = f'boletas_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'

    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename={filename}'

    return response


# =============================================================================
# API: MEDIDORES POR CLIENTE (para filtros dinamicos)
# =============================================================================

@boletas_bp.route('/api/medidores/<int:cliente_id>')
@admin_required
def api_medidores(cliente_id):
    """Retorna medidores de un cliente en formato JSON."""
    from flask import jsonify
    medidores = listar_medidores(cliente_id)
    return jsonify([dict(m) for m in medidores])


# =============================================================================
# HELPER: PRESERVACION DE FILTROS
# =============================================================================

def construir_url_con_filtros(endpoint):
    """Construye URL preservando filtros y ordenamiento del formulario."""
    filtros = {}

    # Capturar filtros del formulario POST
    if request.method == 'POST':
        filtros = {
            'cliente_id': request.form.get('cliente_id'),
            'medidor_id': request.form.get('medidor_id'),
            'año': request.form.get('año'),
            'mes': request.form.get('mes'),
            'pagada': request.form.get('pagada'),
            'sin_comprobante': request.form.get('sin_comprobante'),
            'sort_by': request.form.get('sort_by'),
            'sort_order': request.form.get('sort_order')
        }

    # Limpiar valores vacíos
    filtros = {k: v for k, v in filtros.items() if v}

    return url_for(endpoint, **filtros)


# =============================================================================
# NUEVO SISTEMA DE PAGOS
# =============================================================================

@boletas_bp.route('/pagos')
@admin_required
def pagos_lista():
    """Lista todos los pagos registrados (nuevo historial unificado)."""
    estado = request.args.get('estado')
    cliente_id = request.args.get('cliente_id', type=int)

    pagos = listar_pagos(cliente_id=cliente_id, estado=estado)
    clientes = listar_clientes()

    return render_template('boletas/pagos_lista.html',
                          pagos=pagos,
                          clientes=clientes,
                          filtros={
                              'estado': estado,
                              'cliente_id': cliente_id
                          })


@boletas_bp.route('/pagos/<int:pago_id>')
@admin_required
def pago_detalle(pago_id):
    """Muestra detalle de un pago con sus boletas asociadas."""
    pago = obtener_pago(pago_id)
    if not pago:
        flash('Pago no encontrado', 'error')
        return redirect(url_for('boletas.pagos_lista'))

    return render_template('boletas/pago_detalle.html', pago=pago)


@boletas_bp.route('/pagos/<int:pago_id>/aprobar', methods=['POST'])
@admin_required
def aprobar_pago_route(pago_id):
    """Aprueba un pago en revisión."""
    usuario = get_current_user()
    usuario_id = usuario['id'] if usuario else None

    exito, mensaje = aprobar_pago(pago_id, usuario_id)

    if exito:
        flash(mensaje, 'success')
    else:
        flash(mensaje, 'error')

    return redirect(url_for('boletas.pagos_lista', estado='en_revision'))


@boletas_bp.route('/pagos/<int:pago_id>/rechazar', methods=['POST'])
@admin_required
def rechazar_pago_route(pago_id):
    """Rechaza un pago en revisión."""
    motivo = request.form.get('motivo', '').strip()
    if not motivo:
        flash('Debe proporcionar un motivo de rechazo', 'error')
        return redirect(url_for('boletas.pagos_lista', estado='en_revision'))

    usuario = get_current_user()
    usuario_id = usuario['id'] if usuario else None

    exito, mensaje = rechazar_pago(pago_id, motivo, usuario_id)

    if exito:
        flash(mensaje, 'success')
    else:
        flash(mensaje, 'error')

    return redirect(url_for('boletas.pagos_lista', estado='en_revision'))


@boletas_bp.route('/registrar-pago', methods=['GET', 'POST'])
@admin_required
def registrar_pago_admin():
    """Registra un pago directo desde el admin (ej: pago en efectivo)."""
    from datetime import datetime
    from decimal import Decimal

    if request.method == 'POST':
        cliente_id = request.form.get('cliente_id', type=int)
        monto = request.form.get('monto', type=float)
        metodo_pago = request.form.get('metodo_pago', 'efectivo')
        boletas_ids = request.form.getlist('boletas')
        fecha_pago_str = request.form.get('fecha_pago')
        notas = request.form.get('notas', '').strip() or None

        if not cliente_id or not monto or not boletas_ids:
            flash('Debe seleccionar cliente, monto y al menos una boleta', 'error')
            return redirect(url_for('boletas.registrar_pago_admin'))

        try:
            fecha_pago = datetime.strptime(fecha_pago_str, '%Y-%m-%d').date() if fecha_pago_str else None
        except ValueError:
            fecha_pago = None

        usuario = get_current_user()
        usuario_id = usuario['id'] if usuario else None

        try:
            boletas_ids = [int(b) for b in boletas_ids]
            resultado = registrar_pago_directo(
                cliente_id=cliente_id,
                monto_total=Decimal(str(monto)),
                boletas_ids=boletas_ids,
                metodo_pago=metodo_pago,
                usuario_id=usuario_id,
                fecha_pago=fecha_pago,
                notas=notas
            )

            msg = f'Pago {resultado["numero_pago"]} registrado. '
            msg += f'${resultado["monto_aplicado"]:,.0f} aplicado a {len(resultado["boletas_afectadas"])} boleta(s).'
            if resultado['saldo_generado'] > 0:
                msg += f' ${resultado["saldo_generado"]:,.0f} agregado a saldo a favor.'

            flash(msg, 'success')
            return redirect(url_for('boletas.pago_detalle', pago_id=resultado['pago_id']))

        except Exception as e:
            flash(f'Error al registrar pago: {str(e)}', 'error')
            return redirect(url_for('boletas.registrar_pago_admin'))

    # GET: Mostrar formulario
    cliente_id = request.args.get('cliente_id', type=int)
    clientes = listar_clientes()
    today = datetime.now().strftime('%Y-%m-%d')

    # Obtener boletas pendientes del cliente seleccionado
    boletas_pendientes = []
    if cliente_id:
        boletas_pendientes = listar_boletas(cliente_id=cliente_id, pagada=0)

    return render_template('boletas/registrar_pago.html',
                          clientes=clientes,
                          boletas_pendientes=boletas_pendientes,
                          cliente_id=cliente_id,
                          today=today)


# =============================================================================
# SALDOS DE CLIENTES
# =============================================================================

@boletas_bp.route('/saldos')
@admin_required
def saldos_lista():
    """Lista todos los clientes con su saldo a favor."""
    saldos = listar_saldos_clientes()
    return render_template('boletas/saldos_lista.html', saldos=saldos)


@boletas_bp.route('/saldos/<int:cliente_id>')
@admin_required
def saldo_detalle(cliente_id):
    """Muestra detalle del saldo de un cliente."""
    from src.models import obtener_cliente
    from src.models_pagos import obtener_historial_movimientos

    cliente = obtener_cliente(cliente_id)
    if not cliente:
        flash('Cliente no encontrado', 'error')
        return redirect(url_for('boletas.saldos_lista'))

    resumen = obtener_resumen_cuenta_cliente(cliente_id)
    movimientos = obtener_historial_movimientos(cliente_id)

    return render_template('boletas/saldo_detalle.html',
                          cliente=cliente,
                          resumen=resumen,
                          movimientos=movimientos)


@boletas_bp.route('/saldos/<int:cliente_id>/ajustar', methods=['POST'])
@admin_required
def ajustar_saldo(cliente_id):
    """Realiza un ajuste manual al saldo del cliente."""
    from decimal import Decimal

    monto = request.form.get('monto', type=float)
    descripcion = request.form.get('descripcion', '').strip()

    if monto is None or not descripcion:
        flash('Debe proporcionar monto y descripción', 'error')
        return redirect(url_for('boletas.saldo_detalle', cliente_id=cliente_id))

    usuario = get_current_user()
    usuario_id = usuario['id'] if usuario else None

    exito, mensaje = ajustar_saldo_cliente(
        cliente_id=cliente_id,
        monto=Decimal(str(monto)),
        descripcion=descripcion,
        usuario_id=usuario_id
    )

    if exito:
        flash(mensaje, 'success')
    else:
        flash(mensaje, 'error')

    return redirect(url_for('boletas.saldo_detalle', cliente_id=cliente_id))


# =============================================================================
# API: BOLETAS PENDIENTES POR CLIENTE
# =============================================================================

@boletas_bp.route('/api/boletas-pendientes/<int:cliente_id>')
@admin_required
def api_boletas_pendientes(cliente_id):
    """Retorna boletas pendientes de un cliente en formato JSON."""
    from flask import jsonify
    boletas = listar_boletas(cliente_id=cliente_id, pagada=0)
    return jsonify([{
        'id': b['id'],
        'numero_boleta': b['numero_boleta'],
        'periodo': f"{b['periodo_mes']}/{b['periodo_año']}",
        'total': float(b['total']),
        'saldo_pendiente': float(b.get('saldo_pendiente', b['total']))
    } for b in boletas])
